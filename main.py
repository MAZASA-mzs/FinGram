import asyncio
import sqlite3
import re
import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, Set

# --- Aiogram (Бот) ---
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, BaseFilter
from aiogram.types import Message, Document, CallbackQuery, FSInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder

# --- Конфигурация ---
# ⚠️ Вставьте сюда ваш токен. В реальном проекте - из env.
BOT_TOKEN = "YOUR_TELEGRAM_BOT_TOKEN_HERE"
DB_NAME = "fingram_mvp.db"
# ⚠️ Добавьте сюда ваш Telegram User ID для доступа
ALLOWED_USERS_IDS: Set[int] = {123456789}

# --- Фильтр доступа (White-list) ---
class IsAllowed(BaseFilter):
    """
    Проверяет, есть ли ID пользователя в списке ALLOWED_USERS_IDS.
    """
    async def __call__(self, message: Message) -> bool:
        if message.from_user.id not in ALLOWED_USERS_IDS:
            await message.answer("❌ Доступ запрещен. Это приватный бот.")
            return False
        return True

# --- Модуль 1: База данных (SQLite) ---
def init_database():
    """Инициализирует базу данных и таблицы."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.cursor()
        # Таблица категорий
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
        """)
        # Таблица транзакций (из .xlsx)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            trans_date TEXT NOT NULL,
            amount REAL NOT NULL,
            description TEXT,
            is_matched BOOLEAN DEFAULT 0,
            category_id INTEGER,
            user_note TEXT,
            FOREIGN KEY (category_id) REFERENCES categories (id)
        )
        """)
        # Добавим пару категорий по умолчанию для теста
        try:
            cursor.executemany("INSERT INTO categories (name) VALUES (?)",
                               [('Продукты',), ('Транспорт',), ('Развлечения',), ('Другое',)])
        except sqlite3.IntegrityError:
            pass # Категории уже существуют
        conn.commit()

# --- Модуль 2: Управление категориями (Бот и API) ---
async def db_add_category(name: str) -> bool:
    """Добавляет новую категорию в БД."""
    with sqlite3.connect(DB_NAME) as conn:
        try:
            conn.execute("INSERT INTO categories (name) VALUES (?)", (name,))
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False # Такая категория уже есть

async def db_get_categories() -> list:
    """Получает список всех категорий из БД."""
    with sqlite3.connect(DB_NAME) as conn:
        cursor = conn.execute("SELECT id, name FROM categories")
        return cursor.fetchall()

# --- Модуль 3: Парсер .xlsx (openpyxl) ---
async def parse_sber_xlsx(file_path: str) -> (int, int):
    """
    Парсит .xlsx файл (формат Сбербанка).
    [cite: 36]
    ⚠️ ВАЖНО: Мы делаем допущение о структуре файла.
    Например:
    - Дата в колонке A (начиная с 9 строки)
    - Описание в колонке C
    - Сумма (расход) в колонке F
    Вам нужно будет адаптировать `col_idx` под ваш реальный файл.
    """
    from openpyxl import load_workbook

    # ⚠️ Адаптируйте эти индексы (A=1, B=2, C=3...)
    DATE_COL = 1
    DESC_COL = 3
    AMOUNT_COL = 6 # Расходы
    START_ROW = 9  # Обычно данные начинаются не с первой строки

    transactions_to_add = []
    processed_count = 0
    skipped_count = 0

    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=START_ROW, values_only=True):
            if not row[DATE_COL] or not row[AMOUNT_COL]:
                skipped_count += 1
                continue

            try:
                # '28.09.2025' -> datetime object
                trans_date_str = str(row[DATE_COL]).split(" ")[0]
                trans_date = datetime.strptime(trans_date_str, '%d.%m.%Y').isoformat()

                # '1 500,00' -> -1500.0 (Сбербанк ставит расходы с минусом)
                amount_str = str(row[AMOUNT_COL]).replace(' ', '').replace(',', '.')
                amount = float(amount_str)

                # Нас интересуют только расходы (отрицательные значения)
                if amount < 0:
                    description = str(row[DESC_COL] or "Без описания")
                    transactions_to_add.append((trans_date, abs(amount), description))
                    processed_count += 1
                else:
                    skipped_count += 1 # Это пополнение, пропускаем

            except (ValueError, TypeError) as e:
                print(f"Ошибка парсинга строки: {e}")
                skipped_count += 1

        # Запись в БД
        if transactions_to_add:
            with sqlite3.connect(DB_NAME) as conn:
                conn.executemany(
                    "INSERT INTO transactions (trans_date, amount, description) VALUES (?, ?, ?)",
                    transactions_to_add
                )
                conn.commit()

    except Exception as e:
        print(f"Ошибка чтения .xlsx: {e}")
        return 0, 0
    finally:
        os.remove(file_path) # Удаляем временный файл

    return processed_count, skipped_count

# --- Модуль 4: Matching Engine (Гибридный) ---

def parse_user_note(text: str) -> Optional[tuple[float, str]]:
    """
    Извлекает сумму и текст из заметки пользователя.
    "Такси 500" -> (500.0, "Такси")
    "Купил продукты 1250.50" -> (1250.50, "Купил продукты")
    """
    # Ищет число (целое или с точкой/запятой) в конце строки
    match = re.search(r'(\d+([.,]\d{1,2})?)$', text)
    if not match:
        # Ищет число в начале строки
        match = re.search(r'^(\d+([.,]\d{1,2})?)\b', text)

    if match:
        amount_str = match.group(1).replace(',', '.')
        amount = float(amount_str)
        # Текст - это всё, что не является суммой
        description = text.replace(match.group(0), '').strip()
        return amount, description or "Заметка"

    return None

async def db_find_unmatched(amount: float, days_window: int = 1) -> Optional[dict]:
    """
    Ищет в БД несопоставленную транзакцию по сумме и времени.
    """
    time_limit = (datetime.now() - timedelta(days=days_window)).isoformat()

    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row # Возвращать как dict
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT id, trans_date, description, amount 
            FROM transactions
            WHERE is_matched = 0 AND amount = ? AND trans_date >= ?
            ORDER BY trans_date DESC
            LIMIT 1
            """,
            (amount, time_limit)
        )
        result = cursor.fetchone()
        return dict(result) if result else None

async def db_match_transaction(trans_id: int, category_id: int, user_note: str):
    """Обновляет транзакцию, помечая ее как сопоставленную."""
    with sqlite3.connect(DB_NAME) as conn:
        conn.execute(
            """
            UPDATE transactions
            SET is_matched = 1, category_id = ?, user_note = ?
            WHERE id = ?
            """,
            (category_id, user_note, trans_id)
        )
        conn.commit()

# --- Модуль 5: Генератор Отчётов (pandas -> .csv)  ---
async def generate_csv_report() -> Optional[str]:
    """
    Генерирует CSV-отчет по сопоставленным транзакциям[cite: 41].
    """
    try:
        with sqlite3.connect(DB_NAME) as conn:
            query = """
            SELECT 
                t.trans_date,
                t.amount,
                c.name as category,
                t.user_note,
                t.description as bank_description
            FROM transactions t
            JOIN categories c ON t.category_id = c.id
            WHERE t.is_matched = 1
            ORDER BY t.trans_date
            """
            df = pd.read_sql_query(query, conn)

        if df.empty:
            return None

        report_path = "report_fingram.csv"
        # 'utf-8-sig' для корректного отображения в Excel
        df.to_csv(report_path, index=False, encoding='utf-8-sig')
        return report_path

    except Exception as e:
        print(f"Ошибка генерации отчета: {e}")
        return None

# --- Инициализация Бота и Диспетчера ---
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# --- Хэндлеры (Обработчики команд) ---

@dp.message(IsAllowed(), Command("start"))
async def cmd_start(message: Message):
    await message.answer(
        "👋 Привет! Я FinGram - ваш локальный финансовый помощник.\n"
        "Отправьте мне .xlsx выписку из банка для начала работы.\n"
		"Затем просто пишите мне свои траты, например: `Такси 500`\n\n"
        "Команды:\n"
        "/upload - Загрузить .xlsx выписку\n"
        "/report - Получить .csv отчет [cite: 40]\n"
        "/add_category <Название> - Добавить категорию\n"
        "/list_categories - Показать все категории"
    )

# --- Блок управления категориями ---
@dp.message(IsAllowed(), Command("add_category"))
async def cmd_add_category(message: Message):
    category_name = message.text.replace("/add_category", "").strip()
    if not category_name:
        await message.answer("Ошибка. Укажите название, например:\n`/add_category Еда`")
        return

    if await db_add_category(category_name):
        await message.answer(f"✅ Категория '{category_name}' добавлена.")
    else:
        await message.answer(f"⚠️ Категория '{category_name}' уже существует.")

@dp.message(IsAllowed(), Command("list_categories"))
async def cmd_list_categories(message: Message):
    categories = await db_get_categories()
    if not categories:
        await message.answer("Категории еще не добавлены. Используйте `/add_category`")
        return

    text = "Спиcок ваших категорий:\n" + "\n".join([f"• {name}" for id, name in categories])
    await message.answer(text)

# --- Блок загрузки .xlsx ---
@dp.message(IsAllowed(), Command("upload"))
async def cmd_upload_prompt(message: Message):
    await message.answer("Пожалуйста, прикрепите и отправьте ваш .xlsx файл[cite: 36].")

@dp.message(IsAllowed(), F.document)
async def handle_document(message: Message, bot: Bot):
    document = message.document
    if not document.file_name.endswith('.xlsx'):
        await message.answer("⚠️ Ошибка. Пожалуйста, отправьте файл в формате `.xlsx`.")
        return

    await message.answer(f"⏳ Начинаю обработку файла `{document.file_name}`...")

    # Скачиваем файл
    file_info = await bot.get_file(document.file_id)
    file_path = f"temp_{document.file_id}.xlsx"
    await bot.download_file(file_info.file_path, destination=file_path)

    # Парсим [cite: 37]
    processed, skipped = await parse_sber_xlsx(file_path)

    if processed == 0 and skipped == 0:
        await message.answer("❌ Не удалось прочитать файл. Проверьте формат. [cite: 50]")
    else:
        await message.answer(
            f"✅ Файл обработан.\n"
            f"• Загружено новых транзакций (расходов): {processed}\n"
            f"• Пропущено (пополнения/ошибки): {skipped}"
        )

# --- Блок Matching Engine (текстовые сообщения) ---
@dp.message(IsAllowed(), F.text)
async def handle_user_note(message: Message):
    parsed_data = parse_user_note(message.text)

    if not parsed_data:
        await message.answer("Не могу распознать сумму в сообщении. Попробуйте формат: `Текст Сумма`, например: `Такси 500`")
        return

    amount, description = parsed_data

    # Ищем транзакцию
    found_trans = await db_find_unmatched(amount)

    if not found_trans:
        await message.answer(f"Заметку '`{message.text}`' сохранил, но не нашел подходящей транзакции на {amount} руб. за последние 24ч.")
        # TODO: В будущем - сохранить заметку в отдельную таблицу user_notes
        return

    # Транзакция найдена! Предлагаем категории.
    categories = await db_get_categories()
    if not categories:
        await message.answer("Нашел транзакцию, но у вас нет категорий. Добавьте их: `/add_category <Название>`")
        return

    builder = InlineKeyboardBuilder()
    for cat_id, cat_name in categories:
        # Формируем callback_data: "match_ТранзакцияID_КатегорияID_ТекстЗаметки"
        # Текст заметки кодируем, т.к. он может быть длинным (но для MVP OK)
        callback_data = f"match_{found_trans['id']}_{cat_id}_{message.text[:30]}"
        builder.button(text=cat_name, callback_data=callback_data)

    builder.adjust(2) # По 2 кнопки в ряд

    await message.answer(
        f"Найдена транзакция:\n"
        f"• *{found_trans['description']}* на *{found_trans['amount']} руб.*\n\n"
        f"К какой категории отнести заметку '`{message.text}`'?",
        reply_markup=builder.as_markup(),
        parse_mode="Markdown"
    )

# --- Блок Callback (нажатия кнопок) ---
@dp.callback_query(IsAllowed(), F.data.startswith("match_"))
async def cq_handle_match(callback: CallbackQuery):
    # "match_ТранзакцияID_КатегорияID_ТекстЗаметки"
    try:
        _, trans_id, cat_id, user_note = callback.data.split("_", 3)
        trans_id = int(trans_id)
        cat_id = int(cat_id)
    except ValueError:
        await callback.answer("Ошибка! Неверный формат callback.", show_alert=True)
        return

    # Сопоставляем в БД
    await db_match_transaction(trans_id, cat_id, user_note)

    # Отвечаем пользователю и убираем кнопки
    category_name = [c[1] for c in await db_get_categories() if c[0] == cat_id][0]

    await callback.message.edit_text(
        f"✅ *Сопоставлено:*\n"
        f"Заметка: `{user_note}`\n"
        f"Категория: *{category_name}*",
        parse_mode="Markdown",
        reply_markup=None # Убираем клавиатуру
    )
    await callback.answer(f"Сохранено в '{category_name}'")


# --- Блок отчетов [cite: 40] ---
@dp.message(IsAllowed(), Command("report"))
async def cmd_report(message: Message):
    await message.answer("⏳ Готовлю .csv отчет...")

    report_file_path = await generate_csv_report()

    if not report_file_path:
        await message.answer("Нет данных для отчета. Сначала сопоставьте несколько транзакций.")
        return

    # Отправляем файл [cite: 41]
    csv_file = FSInputFile(report_file_path)
    await message.answer_document(csv_file, caption="Ваш .csv отчет по категориям.")

    # Удаляем временный файл отчета
    os.remove(report_file_path)


# --- Функция запуска ---
async def main():
    print("Инициализация базы данных...")
    init_database()

    print("Запуск бота...")
    # 'IsAllowed()' в 'dp.message' и 'dp.callback_query' применяет фильтр ко всем хэндлерам
    dp.message.filter(IsAllowed())
    dp.callback_query.filter(IsAllowed())

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    # Проверка токена
    if BOT_TOKEN == "YOUR_TELEGRAM_BOT_TOKEN_HERE":
        print("!!! ОШИБКА !!!")
        print("Пожалуйста, укажите ваш BOT_TOKEN в переменной BOT_TOKEN на строке 20.")
    elif not ALLOWED_USERS_IDS or ALLOWED_USERS_IDS == {123456789}:
         print("!!! ВНИМАНИЕ !!!")
         print("Пожалуйста, укажите ваш Telegram User ID в переменной ALLOWED_USERS_IDS на строке 23.")
         print("Вы можете узнать свой ID у бота @userinfobot")
         print("Запуск через 5 секунд...")
         asyncio.run(main())
    else:
        asyncio.run(main())