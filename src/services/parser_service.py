import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any, Optional
import datetime
import logging
import io
from src.models import Transaction # Импортируем модель, но возвращаем dict

log = logging.getLogger(__name__)

# TODO: Уточнить реальную структуру Сбербанк XLSX.
# Это ПРЕДПОЛОЖИТЕЛЬНАЯ структура. Ее нужно будет
# адаптировать под реальный файл.
# (Дата, Время, Сумма, Описание)

EXPECTED_COLUMNS = {
    'Дата операции': 'date',
    'Время операции': 'time',
    'Сумма операции': 'amount',
    'Описание': 'description'
}

def find_header_row(sheet: Worksheet) -> Optional[int]:
    """
    Ищет строку, которая содержит заголовки.
    """
    expected_keys = set(EXPECTED_COLUMNS.keys())
    for i, row in enumerate(sheet.iter_rows(max_rows=20), 1):
        row_values = {str(cell.value).strip() for cell in row if cell.value}
        if expected_keys.issubset(row_values):
            log.info(f"Найдена строка заголовков: {i}")
            return i
    return None

def parse_sberbank_xlsx(file_content: bytes, user_id: int) -> List[Dict[str, Any]]:
    """
    Парсит XLSX файл выписки Сбербанка (предположительный формат).
    Возвращает список словарей, готовых для создания Transaction.
    """
    transactions = []
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        sheet = workbook.active

        header_row_index = find_header_row(sheet)
        if header_row_index is None:
            raise ValueError("Не удалось найти строку заголовков в XLSX файле.")

        headers = {} # 'Дата операции': 0, 'Время операции': 1, ...
        for i, cell in enumerate(sheet[header_row_index]):
            if cell.value in EXPECTED_COLUMNS:
                headers[EXPECTED_COLUMNS[cell.value]] = i

        if len(headers) != len(EXPECTED_COLUMNS):
            raise ValueError(f"Отсутствуют необходимые колонки. Найдено: {headers.keys()}")

        row_count = 0
        for row in sheet.iter_rows(min_row=header_row_index + 1):
            if not row[headers['date']].value:
                # Пустая строка, конец данных
                break

            row_count += 1
            try:
                # 1. Парсим дату и время
                # openpyxl может вернуть datetime.datetime или строку
                op_date_val = row[headers['date']].value
                op_time_val = row[headers['time']].value

                op_date = None
                op_time = None

                if isinstance(op_date_val, datetime.datetime):
                    op_date = op_date_val.date()
                elif isinstance(op_date_val, str):
                    op_date = datetime.datetime.strptime(op_date_val, "%d.%m.%Y").date()

                if isinstance(op_time_val, datetime.time):
                    op_time = op_time_val
                elif isinstance(op_time_val, str):
                    op_time = datetime.datetime.strptime(op_time_val, "%H:%M:%S").time()
                # Иногда время может быть в datetime
                elif isinstance(op_time_val, datetime.datetime):
                    op_time = op_time_val.time()

                if not op_date or not op_time:
                    log.warning(f"Строка {header_row_index + row_count}: Не удалось распознать дату/время. Пропуск.")
                    continue

                timestamp = datetime.datetime.combine(op_date, op_time)

                # 2. Парсим сумму
                # Сумма может быть строкой "1 000,00" или числом
                amount_str = str(row[headers['amount']].value).replace(',', '.').replace(' ', '')
                amount = float(amount_str)

                # Нас интересуют только расходы (отрицательные)
                if amount >= 0:
                    continue

                # 3. Описание
                description = str(row[headers['description']].value)

                transactions.append({
                    "user_id": user_id,
                    "timestamp": timestamp,
                    "amount": amount,
                    "description": description
                })

            except (ValueError, TypeError) as e:
                log.warning(f"Строка {header_row_index + row_count}: Ошибка парсинга: {e}. Пропуск.")

        log.info(f"Успешно спарсено {len(transactions)} транзакций.")
        return transactions

    except Exception as e:
        log.error(f"Ошибка парсинга XLSX файла: {e}", exc_info=True)
        # Возвращаем пустой список или пробрасываем ошибку
        raise ValueError(f"Ошибка обработки файла: {e}")
